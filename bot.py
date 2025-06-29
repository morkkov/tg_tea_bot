from aiogram import Bot, Dispatcher, types, F
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.utils.keyboard import ReplyKeyboardBuilder
from aiogram.client.default import DefaultBotProperties
import asyncio
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
import openpyxl
import os
from datetime import datetime
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.state import State, StatesGroup

from dotenv import load_dotenv
load_dotenv()
import os

API_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "935773354"))

if not API_TOKEN:
    raise ValueError("BOT_TOKEN не найден в .env файле")


class OrderForm(StatesGroup):
    fill_address = State()


class DeliveryInfo(StatesGroup):
    choosing_delivery = State()
    entering_data = State()


class AdminStates(StatesGroup):
    waiting_for_broadcast_message = State()


bot = Bot(token=API_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

cart = {}
# Список всех пользователей бота
users = set()

# Функции для работы с пользователями
def save_users():
    """Сохраняет список пользователей в файл"""
    print(f"Сохраняю пользователей в файл: {users}")
    with open("users.txt", "w", encoding="utf-8") as f:
        for user_id in users:
            f.write(f"{user_id}\n")

def load_users():
    """Загружает список пользователей из файла"""
    try:
        with open("users.txt", "r", encoding="utf-8") as f:
            for line in f:
                user_id = line.strip()
                if user_id.isdigit():
                    users.add(int(user_id))
        print(f"Загружено пользователей из файла: {len(users)}")
    except FileNotFoundError:
        print("Файл users.txt не найден, начинаем с пустого списка")
        pass  # Файл не существует, начинаем с пустого списка

# Загружаем пользователей при запуске
load_users()

from aiogram.fsm.state import State, StatesGroup

class OrderStates(StatesGroup):
    entering_info = State()
    confirming_order = State()
    choosing_delivery = State()

# Загрузка каталога
def load_catalog():
    wb = openpyxl.load_workbook("tea_catalog.xlsx")
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {
            'id': row[0],
            'name': row[1],
            'desc': row[2],
            'photo': row[3],
            'price_100g': row[4],
            'cost_100g': row[5]
        }
        items.append(item)
    return items


import os
import openpyxl
from datetime import datetime

def save_order_to_excel(user_id, items, user_data):
    file_name = "orders.xlsx"
    file_exists = os.path.exists(file_name)

    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Товары", "Сумма", "Прибыль", "Время", "Оплачен", "Данные пользователя", "Детали товаров"])
    else:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

    # Формируем строку с товарами
    names = ", ".join([f"{item['name']} ({item['weight']}г)" for item in items])
    total = sum(item['price'] for item in items)
    profit = sum(item['profit'] for item in items)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Формируем детальную информацию о товарах
    items_details = []
    for item in items:
        items_details.append(f"{item['name']} - {item['weight']}г - {item['price']:.2f}₽")
    items_details_str = "\n".join(items_details)

    # Формируем строку с данными пользователя
    user_info = (
        f"Доставка: {user_data.get('delivery')}\n"
        f"Оплата: {user_data.get('payment_method', 'Не указано')}\n"
        f"Данные: {user_data.get('buyerdata')}\n"
        f"Ник: @{user_data.get('username', 'не указан')}"
    )
    
    # Записываем строку в таблицу
    ws.append([
        user_id,           # A - ID
        names,             # B - Товары (кратко)
        round(total, 2),   # C - Сумма
        round(profit, 2),  # D - Прибыль
        timestamp,         # E - Время
        "Нет",             # F - Оплачен
        user_info,         # G - Данные пользователя
        items_details_str  # H - Детали товаров
    ])

    wb.save(file_name)



@dp.message()
async def handle_all_messages(message: types.Message, state: FSMContext):
    text = message.text
    user_id = message.from_user.id
    current_state = await state.get_state()
    
    print(f"=== ОБРАБОТЧИК handle_all_messages ===")
    print(f"Текст: {text}")
    print(f"Текущее состояние: {current_state}")
    print(f"Ожидаемое состояние: {DeliveryInfo.entering_data}")

    if current_state == OrderStates.entering_info:
        return
    if current_state == DeliveryInfo.entering_data:
        print("Состояние совпадает! Вызываем save_delivery_info")
        await save_delivery_info(message, state)
        return
    if current_state == AdminStates.waiting_for_broadcast_message:
        print("Состояние AdminStates.waiting_for_broadcast_message - передаем в handle_broadcast_message")
        await handle_broadcast_message(message, state)
        return
        
    if text == "/start":
        # Добавляем пользователя в список
        users.add(user_id)
        print(f"Добавлен новый пользователь: {user_id}")
        save_users()  # Сохраняем в файл
        
        kb = ReplyKeyboardBuilder()
        kb.add(types.KeyboardButton(text="📦 Каталог"))
        kb.add(types.KeyboardButton(text="🛒 Корзина"))
        kb.add(types.KeyboardButton(text="Связь с администратором"))
        if user_id == ADMIN_ID:
            kb.add(types.KeyboardButton(text="🔐 Админ-панель"))
        kb.adjust(2)
        await message.answer("Добро пожаловать в чайный бот!", reply_markup=kb.as_markup(resize_keyboard=True))

    elif text == "📦 Каталог":
        catalog = load_catalog()
        for item in catalog:
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="25г", callback_data=f"add_{item['id']}_25"),
                    InlineKeyboardButton(text="50г", callback_data=f"add_{item['id']}_50")
                ],
                [
                    InlineKeyboardButton(text="100г", callback_data=f"add_{item['id']}_100"),
                    InlineKeyboardButton(text="200г", callback_data=f"add_{item['id']}_200")
                ]
            ])
            await message.answer_photo(
                photo=item['photo'],
                caption=f"<b>{item['name']}</b>\n\n{item['desc']}\n\n💵 Цена: {item['price_100g']}₽ / 100г",
                reply_markup=kb
            )

    elif text == "🛒 Корзина":
        items = cart.get(user_id, [])
        if not items:
            return await message.answer("Ваша корзина пуста.")
        text_cart = "🧺 Ваша корзина:\n"
        total = 0
        for i, item in enumerate(items, 1):
            text_cart += f"{i}. {item['name']} — {item['weight']}г — {item['price']:.2f}₽\n"
            total += item['price']
        text_cart += f"\n💰 Общая сумма: {total:.2f}₽"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Оформить заказ", callback_data="checkout")],
            [InlineKeyboardButton(text="🗑️ Очистить корзину", callback_data="clear_cart")]
        ])
        await message.answer(text_cart, reply_markup=kb)

        

    elif text == "🔐 Админ-панель" and user_id == ADMIN_ID:
        # Создаем клавиатуру админ-панели правильно с импортированными классами
        admin_kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Заказы")],
                [KeyboardButton(text="📢 Отправить пост всем")],
                [KeyboardButton(text="🧪 Тестовая рассылка")],
                [KeyboardButton(text="Назад")]
            ],
            resize_keyboard=True
        )
        await message.answer(
            f"Добро пожаловать в админ-панель\n\n"
            f"👥 Всего пользователей: {len(users)}",
            reply_markup=admin_kb
        )

    elif text == "Заказы" and user_id == ADMIN_ID:
        if not cart:
            await message.answer("Пока нет заказов.")
        else:
            text_orders = "Все заказы:\n\n"
            for uid, items in cart.items():
                text_orders += f"Пользователь {uid}:\n"
                for i, item in enumerate(items, 1):
                    text_orders += f"  {i}. {item['name']} — {item['weight']}г — {item['price']:.2f}₽\n"
                text_orders += "\n"
            await message.answer(text_orders)

    elif text == "📢 Отправить пост всем" and user_id == ADMIN_ID:
        await message.answer(
            "Введите сообщение, которое будет отправлено всем пользователям бота:\n\n"
            "💡 Поддерживается HTML разметка\n"
            "❌ Для отмены напишите 'отмена'"
        )
        await state.set_state(AdminStates.waiting_for_broadcast_message)

    elif text == "🧪 Тестовая рассылка" and user_id == ADMIN_ID:
        # Отправляем тестовое сообщение только администратору
        try:
            await message.answer("🧪 Тестовое сообщение для проверки работы бота")
            await message.answer(
                f"✅ Тестовая рассылка успешна!\n\n"
                f"📊 Информация:\n"
                f"👥 Всего пользователей в базе: {len(users)}\n"
                f"🆔 Ваш ID: {user_id}\n"
                f"📝 Пользователи: {list(users)[:5]}{'...' if len(users) > 5 else ''}"
            )
        except Exception as e:
            await message.answer(f"❌ Ошибка тестовой рассылки: {e}")

    elif text == "Связь с администратором":
        await message.answer("Для связи с администратором напишите @jdueje")

    elif text == "Назад":
        kb = ReplyKeyboardBuilder()
        kb.add(types.KeyboardButton(text="📦 Каталог"))
        kb.add(types.KeyboardButton(text="🛒 Корзина"))
        kb.add(types.KeyboardButton(text="Связь с администратором"))
        if user_id == ADMIN_ID:
            kb.add(types.KeyboardButton(text="🔐 Админ-панель"))
        kb.adjust(2)
        await message.answer("Главное меню:", reply_markup=kb.as_markup(resize_keyboard=True))




@dp.callback_query(F.data.startswith("add_"))
async def add_to_cart(callback: types.CallbackQuery):
    try:
        _, tea_id_str, weight_str = callback.data.split("_")
        tea_id = int(tea_id_str)
        weight = int(weight_str)
    except:
        return await callback.answer("Некорректные данные!")

    user_id = callback.from_user.id
    catalog = load_catalog()
    tea = next((item for item in catalog if item['id'] == tea_id), None)
    if not tea:
        return await callback.answer("Товар не найден!")

    price = (tea['price_100g'] / 100) * weight
    profit = price - ((tea['cost_100g'] / 100) * weight)
    
    item = {
        'name': tea['name'],
        'weight': weight,
        'price': price,
        'profit': profit
    }
    cart.setdefault(user_id, []).append(item)
    await callback.answer(f"Добавлено в корзину: {weight}г {tea['name']}")
    #user_data = {"name": "не указано", "address": "не указано", "delivery_type": "не указано", "username": callback.from_user.username}
    #save_order_to_excel(user_id, cart[user_id], user_data)
        # Сохраняем заказ в Excel (перезаписывается каждый раз — упрощённо)
    


@dp.callback_query(F.data == "checkout")
async def start_checkout(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "Выберите способ доставки:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Европочта +4р", callback_data="delivery_euro")],
            [InlineKeyboardButton(text="Белпочта +3р", callback_data="delivery_bel")],
            [InlineKeyboardButton(text="Сдек", callback_data="delivery_sdek")]
        ])
    )
    await state.set_state(DeliveryInfo.choosing_delivery)

@dp.callback_query(F.data == "clear_cart")
async def clear_cart(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    cart[user_id] = []
    await callback.answer("🗑️ Корзина очищена!")
    await callback.message.answer("🧺 Корзина пуста.")

# ДОБАВЛЕНИЕ: Сохранение выбора и запрос данных
# Убираем текстовый обработчик - оставляем только через кнопки

@dp.message(DeliveryInfo.entering_data)
async def save_delivery_info(message: types.Message, state: FSMContext):
    print(f"=== ОБРАБОТЧИК save_delivery_info СРАБОТАЛ ===")
    print(f"Получено сообщение: {message.text}")
    
    try:
        user_data = await state.get_data()
        print(f"Данные состояния: {user_data}")
        
        if "delivery" not in user_data:
            print("ОШИБКА: delivery не найден в состоянии!")
            await message.answer("❌ Ошибка: способ доставки не выбран. Попробуйте снова.")
            return
            
        delivery_type = user_data["delivery"]
        delivery_info = message.text.strip()
        user_id = message.from_user.id

        user_name = message.from_user.username or "не указан"
        await state.update_data(buyerdata=delivery_info, username=user_name)

        # Сохраняем в Excel
        now = datetime.now().strftime("%d.%m.%Y %H:%M")
        file_name = "delivery_data.xlsx"
        
        print(f"Сохраняем в файл: {file_name}")
        
        if not os.path.exists(file_name):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["User ID", "Username", "Способ доставки", "Данные", "Время", "Заказ"])
            print("Создан новый файл Excel")
        else:
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
            print("Загружен существующий файл Excel")

        # Получаем информацию о заказе
        items = cart.get(user_id, [])
        if items:
            order_info = []
            for item in items:
                order_info.append(f"{item['name']} - {item['weight']}г - {item['price']:.2f}₽")
            order_text = "\n".join(order_info)
            total_sum = sum(item['price'] for item in items)
            order_text += f"\n\nИтого: {total_sum:.2f}₽"
        else:
            order_text = "Корзина пуста"

        ws.append([user_id, user_name, delivery_type, delivery_info, now, order_text])
        wb.save(file_name)
        print(f"Данные сохранены: {user_id}, {user_name}, {delivery_type}, {delivery_info}, {now}")

        # Показываем сообщение об успехе
        await message.answer("✅ Данные успешно сохранены в Excel!")

        # Вычисляем сумму заказа
        items = cart.get(user_id, [])
        total_sum = sum(item['price'] for item in items) if items else 0
        
        # Сохраняем сумму в состоянии
        await state.update_data(order_total=total_sum)

        # Переход к оплате
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💳 Оплатить", callback_data="show_payment")]
        ])
        await message.answer(f"Нажмите кнопку для оплаты (сумма: {total_sum:.2f}₽):", reply_markup=kb)

    except Exception as e:
        print(f"ОШИБКА при сохранении данных: {e}")
        import traceback
        traceback.print_exc()
        await message.answer("❌ Произошла ошибка при сохранении данных. Попробуйте снова.")


@dp.callback_query(F.data.startswith("delivery_"))
async def handle_delivery_choice(callback: types.CallbackQuery, state: FSMContext):
    print(f"=== ОБРАБОТЧИК handle_delivery_choice СРАБОТАЛ ===")
    data = callback.data.replace("delivery_", "")
    delivery_text = {
        "euro": "Европочта",
        "bel": "Белпочта",
        "sdek": "СДЭК"
    }.get(data, "Неизвестно")

    print(f"Выбран способ доставки: {delivery_text}")
    await state.update_data(delivery=delivery_text)
    
    # Создаем клавиатуру с кнопкой "Назад"
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_main")]
    ])
    
    await callback.message.answer(
        f"Введите данные для {delivery_text}:\n\n"
        "📦 ФИО\n🏡 Адрес полностью\n📮 Почтовый индекс\n📱 Номер телефона",
        reply_markup=kb
    )
    
    print(f"Устанавливаем состояние: {DeliveryInfo.entering_data}")
    await state.set_state(DeliveryInfo.entering_data)
    await callback.answer()


@dp.callback_query(F.data == "back_to_main")
async def back_to_main_menu(callback: types.CallbackQuery, state: FSMContext):
    # Очищаем состояние
    await state.clear()
    
    # Создаем главное меню
    kb = ReplyKeyboardBuilder()
    kb.add(types.KeyboardButton(text="📦 Каталог"))
    kb.add(types.KeyboardButton(text="🛒 Корзина"))
    kb.add(types.KeyboardButton(text="Связь с администратором"))
    if callback.from_user.id == ADMIN_ID:
        kb.add(types.KeyboardButton(text="🔐 Админ-панель"))
    kb.adjust(2)
    
    await callback.message.answer("Главное меню:", reply_markup=kb.as_markup(resize_keyboard=True))
    await callback.answer()

@dp.callback_query(F.data == "show_payment")
async def show_payment_info(callback: types.CallbackQuery, state: FSMContext):
    # Генерируем ID заказа
    import random
    order_id = f"ORDER-{random.randint(10000, 99999)}"
    
    # Получаем данные состояния
    user_data = await state.get_data()
    order_total = user_data.get("order_total", 0)
    
    # Сохраняем ID заказа в состоянии
    await state.update_data(order_id=order_id)
    
    # Показываем номер карты
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Я оплатил", callback_data="confirm_payment")]
    ])
    
    await callback.message.answer(
        f"💳 <b>Оплата заказа {order_id}</b>\n\n"
        f"Номер карты: <code>1234 5678 9012 3456</code>\n"
        f"Сумма к оплате: <b>{order_total:.2f}₽</b>\n\n"
        f"После оплаты нажмите кнопку ниже и сделайте скриншот оплаты:",
        reply_markup=kb,
        parse_mode="HTML"
    )
    await callback.answer()

@dp.callback_query(F.data == "confirm_payment")
async def confirm_payment(callback: types.CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    order_id = user_data.get("order_id", "UNKNOWN")
    
    # Сохраняем заказ в основной Excel
    user_id = callback.from_user.id
    items = cart.get(user_id, [])
    
    if items:
        save_order_to_excel(user_id, items, user_data)
    
    # Очищаем корзину
    cart[user_id] = []
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 В главное меню", callback_data="back_to_main")]
    ])
    
    await callback.message.answer(
        f"🎉 <b>Заказ успешно оформлен!</b>\n\n"
        f"ID заказа: <code>{order_id}</code>\n"
        f"Способ доставки: {user_data.get('delivery', 'Не указан')}\n\n"
        f"Мы свяжемся с вами в ближайшее время для уточнения деталей заказа и проверки оплаты!",
        reply_markup=kb,
        parse_mode="HTML"
    )
    
    await state.clear()
    await callback.answer()

@dp.message(AdminStates.waiting_for_broadcast_message)
async def handle_broadcast_message(message: types.Message, state: FSMContext):
    if message.from_user.id != ADMIN_ID:
        return
    
    if message.text and message.text.lower() == "отмена":
        await state.clear()
        admin_kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Заказы")],
                [KeyboardButton(text="📢 Отправить пост всем")],
                [KeyboardButton(text="🧪 Тестовая рассылка")],
                [KeyboardButton(text="Назад")]
            ],
            resize_keyboard=True
        )
        await message.answer("❌ Рассылка отменена", reply_markup=admin_kb)
        return
    
    # Отправляем сообщение всем пользователям
    success_count = 0
    error_count = 0
    
    print(f"Начинаю рассылку. Всего пользователей: {len(users)}")
    print(f"Список пользователей: {users}")
    
    if len(users) == 0:
        await message.answer("❌ Нет пользователей для рассылки!")
        await state.clear()
        return
    
    await message.answer("📤 Начинаю рассылку...")
    
    for user_id in users:
        try:
            print(f"Отправляю сообщение пользователю {user_id}")
            if message.photo:
                await bot.send_photo(
                    chat_id=user_id,
                    photo=message.photo[-1].file_id,
                    caption=message.caption or "",
                    parse_mode="HTML"
                )
            elif message.video:
                await bot.send_video(
                    chat_id=user_id,
                    video=message.video.file_id,
                    caption=message.caption or "",
                    parse_mode="HTML"
                )
            elif message.document:
                await bot.send_document(
                    chat_id=user_id,
                    document=message.document.file_id,
                    caption=message.caption or "",
                    parse_mode="HTML"
                )
            else:
                # Исправляем ошибку с типом данных
                text_content = message.text or ""
                await bot.send_message(
                    chat_id=user_id,
                    text=text_content,
                    parse_mode="HTML"
                )
            success_count += 1
            print(f"✅ Успешно отправлено пользователю {user_id}")
            await asyncio.sleep(0.05)  # Небольшая задержка чтобы не спамить
        except Exception as e:
            error_count += 1
            error_type = type(e).__name__
            print(f"❌ Ошибка отправки пользователю {user_id}: {error_type} - {e}")
            
            # Проверяем конкретные типы ошибок
            if "Forbidden" in str(e):
                print(f"   Пользователь {user_id} заблокировал бота")
            elif "Chat not found" in str(e):
                print(f"   Чат с пользователем {user_id} не найден")
            elif "User is deactivated" in str(e):
                print(f"   Пользователь {user_id} деактивирован")
            else:
                print(f"   Неизвестная ошибка для пользователя {user_id}")
    
    print(f"Рассылка завершена. Успешно: {success_count}, Ошибок: {error_count}")
    
    # Очищаем состояние
    await state.clear()
    
    # Возвращаем админ-панель
    admin_kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Заказы")],
            [KeyboardButton(text="📢 Отправить пост всем")],
            [KeyboardButton(text="🧪 Тестовая рассылка")],
            [KeyboardButton(text="Назад")]
        ],
        resize_keyboard=True
    )
    
    await message.answer(
        f"✅ Рассылка завершена!\n\n"
        f"📊 Статистика:\n"
        f"✅ Успешно отправлено: {success_count}\n"
        f"❌ Ошибок: {error_count}\n"
        f"👥 Всего пользователей: {len(users)}",
        reply_markup=admin_kb
    )

# Запуск бота
async def main():
    print(f"Запуск бота. Загружено пользователей: {len(users)}")
    print(f"Список пользователей: {users}")
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())



